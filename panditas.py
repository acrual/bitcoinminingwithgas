from datetime import datetime, timedelta
from openpyxl import *
import psycopg2



wb = load_workbook('BitcoinMiningAUTO.xlsm', data_only=True)
# sheet = wb['P&L']
sheet2 = wb['MP2']
sql =  "INSERT INTO cifras (date, cifras) VALUES (%s, %s);"
# sheet['D3'] = 2
conn = psycopg2.connect(host="localhost",database="Back y Front End Python", user="adolfo", password="aePOnW8;0c73)LTb")

def mmbtu(num):
    lista = []
    for i in range(num):
        data = (sheet2.cell(row = 725 + i, column = 1).value, sheet2.cell(row = 725 + i, column = 11).value)
        lista.append(data)
        """cur = conn.cursor()
        cur.execute(sql, data) """
        # commit the changes to the database
        conn.commit()
    return lista






